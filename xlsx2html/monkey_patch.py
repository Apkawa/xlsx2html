# type: ignore
import functools

import openpyxl
from openpyxl import utils
from openpyxl.utils import cell
from openpyxl.worksheet.cell_range import CellRange

old_range_boundaries = cell.range_boundaries

MAX_COL_INDEX = 18278
MAX_ROW_INDEX = 1048576


@functools.wraps(old_range_boundaries)
def new_range_boundaries(range_string):
    """
    patch for handle string like
    A:A - all column
    2:2 - all row
    A2:A - from row 2 to end rows
    B3:3 - from column B to end columns
    """
    from openpyxl.utils.cell import ABSOLUTE_RE

    m = ABSOLUTE_RE.match(range_string)
    if m:
        min_col, min_row, sep, max_col, max_row = m.groups()
        if sep:
            if not max_col:
                max_col = utils.get_column_letter(MAX_COL_INDEX)
            if not max_row:
                max_row = str(MAX_ROW_INDEX)
        range_string = "".join(
            [min_col or utils.get_column_letter(1), min_row or "1", ":", max_col, max_row]
        )

    min_col, min_row, max_col, max_row = old_range_boundaries(range_string)
    return min_col or 1, min_row or 1, max_col or MAX_COL_INDEX, max_row or MAX_ROW_INDEX


new_range_boundaries._monkey = True
cell.range_boundaries = new_range_boundaries
utils.range_boundaries = new_range_boundaries
openpyxl.utils = utils


def CellRange__init__(
    self, range_string=None, min_col=None, min_row=None, max_col=None, max_row=None, title=None
):
    if range_string is not None:
        if "!" in range_string:
            from openpyxl.utils import range_to_tuple

            title, (min_col, min_row, max_col, max_row) = range_to_tuple(range_string)
        else:
            # FIX
            min_col, min_row, max_col, max_row = new_range_boundaries(range_string)

    self.min_col = min_col
    self.min_row = min_row
    self.max_col = max_col
    self.max_row = max_row
    self.title = title

    if min_col > max_col:
        fmt = "{max_col} must be greater than {min_col}"
        raise ValueError(fmt.format(min_col=min_col, max_col=max_col))
    if min_row > max_row:
        fmt = "{max_row} must be greater than {min_row}"
        raise ValueError(fmt.format(min_row=min_row, max_row=max_row))


CellRange.__init__ = CellRange__init__
