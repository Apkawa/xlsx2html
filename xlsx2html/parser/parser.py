from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any, Optional, List, Dict, Tuple, Union

import openpyxl
import six
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl.utils import rows_from_range, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from xlsx2html.compat import OPENPYXL_24
from xlsx2html.parser.cell import CellInfo, Borders
from xlsx2html.parser.image import ImageInfo
from xlsx2html.parser.utils import get_sheet, SheetNameType
from xlsx2html.utils.cell import letter_to_col_index, col_index_to_letter


@dataclass
class Column:
    index: int
    letter: str
    width: float
    hidden: bool = False


@dataclass
class MergedCellInfo:
    colspan: Optional[int] = None
    rowspan: Optional[int] = None
    cells: List[Cell] = field(default_factory=list)


ImageInfoMapType = Dict[Tuple[int, int], List[ImageInfo]]


@dataclass
class ParserResult:
    cols: List[Column]
    rows: List[List[CellInfo]]
    images: ImageInfoMapType


class XLSXParser:
    def __init__(self, filepath: Any, locale: str = "en", parse_formula: bool = False):
        self.locale = locale
        self.wb = openpyxl.load_workbook(filepath, data_only=True)
        self.fb = None
        if parse_formula:
            self.fb = openpyxl.load_workbook(filepath, data_only=False)

    def get_sheet_names(self) -> List[str]:
        return self.wb.sheetnames

    def get_sheet(self, sheet: SheetNameType = None) -> ParserResult:
        ws = get_sheet(self.wb, sheet)
        fs = None
        if self.fb:
            fs = get_sheet(self.fb, sheet)

        merged_cell_map = {}
        if OPENPYXL_24:
            merged_cell_ranges = ws.merged_cell_ranges
            excluded_cells = set(ws.merged_cells)
        else:
            merged_cell_ranges = [
                cell_range.coord for cell_range in ws.merged_cells.ranges
            ]
            excluded_cells = set(
                [
                    cell
                    for cell_range in merged_cell_ranges
                    for rows in rows_from_range(cell_range)
                    for cell in rows
                ]
            )

        for cell_range in merged_cell_ranges:
            cell_range_list = list(ws[cell_range])
            m_cell = cell_range_list[0][0]

            colspan = len(cell_range_list[0])
            rowspan = len(cell_range_list)
            merged_cell_map[m_cell.coordinate] = MergedCellInfo(
                colspan=None if colspan <= 1 else colspan,
                rowspan=None if rowspan <= 1 else rowspan,
                cells=[c for rows in cell_range_list for c in rows],
            )

            excluded_cells.remove(m_cell.coordinate)

        max_col_number = 0

        data_list = []
        for row_i, row in enumerate(ws.iter_rows()):
            data_row: List[CellInfo] = []
            data_list.append(data_row)
            for col_i, cell in enumerate(row):
                row_dim = ws.row_dimensions[cell.row]

                # Remove hidden rows
                if cell.coordinate in excluded_cells or row_dim.hidden:
                    continue

                # TODO remove empty columns
                if col_i > max_col_number:
                    max_col_number = col_i

                height = 19

                if row_dim.customHeight:
                    height = round(row_dim.height, 2)

                f_cell = None
                if fs:
                    f_cell = fs[cell.coordinate]

                cell_data = self.get_cell_data(cell, f_cell)
                cell_data.height = height

                merged_cell_info = merged_cell_map.get(cell.coordinate)
                if merged_cell_info:
                    cell_data.colspan = merged_cell_info.colspan
                    cell_data.rowspan = merged_cell_info.rowspan

                if merged_cell_info:
                    cell_data.border = self.merge_borders(
                        [cell] + merged_cell_info.cells
                    )

                data_row.append(cell_data)

        col_list = self.get_columns(ws, max_col_number)

        # Remove hidden columns
        hidden_columns = {col.index for col in col_list if col.hidden}

        for row in data_list:
            for c in row:
                if c.column in hidden_columns:
                    row.remove(c)

        col_list = [col for col in col_list if col.index not in hidden_columns]

        return ParserResult(rows=data_list, cols=col_list, images=self.get_images(ws))

    @staticmethod
    def get_columns(ws: Worksheet, max_col: int) -> List[Column]:
        col_list = []
        max_col_number = max_col

        column_dimensions = sorted(
            ws.column_dimensions.items(), key=lambda d: column_index_from_string(d[0])
        )

        for col_i, col_dim in column_dimensions:
            if not all([col_dim.min, col_dim.max]):
                continue
            width = 0.89
            if col_dim.customWidth:
                width = round(col_dim.width / 10.0, 2)
            col_width = 96 * width

            index = letter_to_col_index(col_dim.index)
            for i in six.moves.range((col_dim.max - col_dim.min) + 1):
                if max_col_number < 0:
                    break
                max_col_number -= 1
                new_index = index + i
                letter = col_index_to_letter(new_index)
                col_list.append(
                    Column(
                        index=new_index,
                        letter=letter,
                        hidden=col_dim.hidden,
                        width=col_width,
                    )
                )
        return col_list

    @staticmethod
    def get_images(ws: Worksheet) -> ImageInfoMapType:
        images: List[Image] = ws._images

        images_data = defaultdict(list)
        for _i in images:
            _id = ImageInfo.from_ws_image(_i)
            images_data[(_id.col, _id.row)].append(_id)
        return images_data

    def get_cell_data(self, cell: Cell, f_cell: Optional[Cell] = None) -> CellInfo:
        return CellInfo.from_cell(cell, f_cell=f_cell, _locale=self.locale)

    @staticmethod
    def merge_borders(cells: List[Cell]) -> Union[Borders, None]:
        border = None
        for m_cell in cells:
            m_border = CellInfo.get_border(m_cell)
            if not m_border:
                continue
            if border is None:
                border = m_border
                continue
            for b_dir in ["right", "left", "top", "bottom"]:
                if not getattr(border, b_dir):
                    setattr(border, b_dir, getattr(m_border, b_dir))
        return border
