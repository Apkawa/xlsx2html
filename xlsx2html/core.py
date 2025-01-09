import contextlib
import io
from collections import defaultdict
from typing import List

import openpyxl
import six
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
from openpyxl.styles.colors import COLOR_INDEX, aRGB_REGEX
from openpyxl.utils import rows_from_range, column_index_from_string, units
from openpyxl.utils.escape import unescape
from openpyxl.worksheet.worksheet import Worksheet

from xlsx2html.compat import OPENPYXL_24
from xlsx2html.constants.border import DEFAULT_BORDER_STYLE, BORDER_STYLES
from xlsx2html.format import format_cell
from xlsx2html.utils.image import bytes_to_datauri


def render_attrs(attrs):
    if not attrs:
        return ""
    return " ".join(
        ['%s="%s"' % a for a in sorted(attrs.items(), key=lambda a: a[0]) if a[1]]
    )


def render_inline_styles(styles):
    if not styles:
        return ""
    return ";".join(
        [
            "%s: %s" % a
            for a in sorted(styles.items(), key=lambda a: a[0])
            if a[1] is not None
        ]
    )


def normalize_color(color):
    # TODO RGBA
    rgb = None
    if color.type == "rgb":
        rgb = color.rgb
    if color.type == "indexed":
        try:
            rgb = COLOR_INDEX[color.indexed]
        except IndexError:
            # The indices 64 and 65 are reserved for the system
            # foreground and background colours respectively
            pass
        if not rgb or not aRGB_REGEX.match(rgb):
            # TODO system fg or bg
            rgb = "00000000"
    if rgb:
        return "#" + rgb[2:]
    return None


def get_border_style_from_cell(cell):
    h_styles = {}
    for b_dir in ["right", "left", "top", "bottom"]:
        b_s = getattr(cell.border, b_dir)
        if not b_s:
            continue
        border_style = BORDER_STYLES.get(b_s.style)
        if border_style is None and b_s.style:
            border_style = DEFAULT_BORDER_STYLE

        if not border_style:
            continue

        for k, v in border_style.items():
            h_styles["border-%s-%s" % (b_dir, k)] = v
        if b_s.color:
            h_styles["border-%s-color" % (b_dir)] = normalize_color(b_s.color)

    return h_styles


def get_styles_from_cell(cell, merged_cell_map=None, default_cell_border="none"):
    merged_cell_map = merged_cell_map or {}

    h_styles = {"border-collapse": "collapse"}
    b_styles = get_border_style_from_cell(cell)
    if merged_cell_map:
        # TODO edged_cells
        for m_cell in merged_cell_map["cells"]:
            b_styles.update(get_border_style_from_cell(m_cell))

    for b_dir in ["border-right", "border-left", "border-top", "border-bottom"]:
        style_tag = b_dir + "-style"
        if (b_dir not in b_styles) and (style_tag not in b_styles):
            b_styles[b_dir] = default_cell_border
    h_styles.update(b_styles)

    if cell.alignment.horizontal:
        h_styles["text-align"] = cell.alignment.horizontal
    if cell.alignment.vertical:
        h_styles["vertical-align"] = cell.alignment.vertical

    with contextlib.suppress(AttributeError):
        if cell.fill.patternType == "solid":
            # TODO patternType != 'solid'
            h_styles["background-color"] = normalize_color(cell.fill.fgColor)

    if cell.font:
        h_styles["font-size"] = "%spx" % cell.font.sz
        if cell.font.color:
            h_styles["color"] = normalize_color(cell.font.color)
        if cell.font.b:
            h_styles["font-weight"] = "bold"
        if cell.font.i:
            h_styles["font-style"] = "italic"
        if cell.font.u:
            h_styles["text-decoration"] = "underline"
    return h_styles


def get_cell_id(cell):
    return "{}!{}".format(cell.parent.title, cell.coordinate)


def image_to_data(image: Image) -> dict:
    _from: AnchorMarker = image.anchor._from
    graphicalProperties: GraphicalProperties = image.anchor.pic.graphicalProperties
    transform = graphicalProperties.transform
    # http://officeopenxml.com/drwSp-location.php
    offsetX = units.EMU_to_pixels(_from.colOff)
    offsetY = units.EMU_to_pixels(_from.rowOff)
    # TODO recalculate to relative cell
    data = {
        "col": _from.col + 1,
        "row": _from.row + 1,
        "offset": {"x": offsetX, "y": offsetY},
        "width": units.EMU_to_pixels(transform.ext.width),
        "height": units.EMU_to_pixels(transform.ext.height),
        "src": bytes_to_datauri(image.ref, image.path),
        "style": {
            "margin-left": f"{offsetX}px",
            "margin-top": f"{offsetY}px",
            "position": "absolute",
        },
    }
    return data


def images_to_data(ws: Worksheet):
    images: List[Image] = ws._images

    images_data = defaultdict(list)
    for _i in images:
        _id = image_to_data(_i)
        images_data[(_id["col"], _id["row"])].append(_id)
    return images_data


def worksheet_to_data(ws, locale=None, fs=None, default_cell_border="none"):
    merged_cell_map = {}
    if OPENPYXL_24:
        merged_cell_ranges = ws.merged_cell_ranges
        excluded_cells = set(ws.merged_cells)
    else:
        merged_cell_ranges = [cell_range.coord for cell_range in ws.merged_cells.ranges]
        excluded_cells = set(
            [
                cell
                for cell_range in merged_cell_ranges
                for rows in rows_from_range(cell_range)
                for cell in rows
            ]
        )

    for cell_range in merged_cell_ranges:
        if ":" not in str(cell_range):
            cell_range_list = list(ws[f"{cell_range}:{cell_range}"])
        else:
            cell_range_list = list(ws[cell_range])

        m_cell = cell_range_list[0][0]

        colspan = len(cell_range_list[0])
        rowspan = len(cell_range_list)
        merged_cell_map[m_cell.coordinate] = {
            "attrs": {
                "colspan": None if colspan <= 1 else colspan,
                "rowspan": None if rowspan <= 1 else rowspan,
            },
            "cells": [c for rows in cell_range_list for c in rows],
        }

        excluded_cells.remove(m_cell.coordinate)

    max_col_number = 0

    data_list = []
    for row_i, row in enumerate(ws.iter_rows()):
        data_row = []
        data_list.append(data_row)
        for col_i, cell in enumerate(row):
            row_dim = ws.row_dimensions[cell.row]

            if cell.coordinate in excluded_cells or row_dim.hidden:
                continue

            if col_i > max_col_number:
                max_col_number = col_i

            height = 19

            if row_dim.customHeight:
                height = round(row_dim.height, 2)

            f_cell = None
            if fs:
                f_cell = fs[cell.coordinate]
            value = cell.value
            if isinstance(value, str):
                value = unescape(value)
            cell_data = {
                "column": cell.column,
                "row": cell.row,
                "value": value,
                "formatted_value": format_cell(cell, locale=locale, f_cell=f_cell),
                "attrs": {"id": get_cell_id(cell)},
                "style": {"height": f"{height}pt"},
            }
            merged_cell_info = merged_cell_map.get(cell.coordinate, {})
            if merged_cell_info:
                cell_data["attrs"].update(merged_cell_info["attrs"])
            cell_data["style"].update(
                get_styles_from_cell(cell, merged_cell_info, default_cell_border)
            )
            data_row.append(cell_data)

    col_list = []
    max_col_number += 1

    column_dimensions = sorted(
        ws.column_dimensions.items(), key=lambda d: column_index_from_string(d[0])
    )

    for col_i, col_dim in column_dimensions:
        if not all([col_dim.min, col_dim.max]):
            continue
        width = 0.89
        if col_dim.customWidth:
            width = round(col_dim.width / 10.0, 2)
        col_width = 96 * width

        for _ in six.moves.range((col_dim.max - col_dim.min) + 1):
            max_col_number -= 1
            col_list.append(
                {
                    "index": col_dim.index,
                    "hidden": col_dim.hidden,
                    "style": {"width": "{}px".format(col_width)},
                }
            )
            if max_col_number < 0:
                break
    return {"rows": data_list, "cols": col_list, "images": images_to_data(ws)}


def render_table(data, append_headers, append_lineno):
    html = [
        "<table  "
        'style="border-collapse: collapse" '
        'border="0" '
        'cellspacing="0" '
        'cellpadding="0">'
        "<colgroup>"
    ]
    hidden_columns = set()
    for col in data["cols"]:
        if col["hidden"]:
            hidden_columns.add(col["index"])
        html.append(
            '<col {attrs} style="{styles}">'.format(
                attrs=render_attrs(col.get("attrs")),
                styles=render_inline_styles(col.get("style")),
            )
        )
    html.append("</colgroup>")

    append_headers(data, html)

    for i, row in enumerate(data["rows"]):
        trow = ["<tr>"]
        append_lineno(trow, i)
        for cell in row:
            if cell["column"] in hidden_columns:
                continue
            images = data["images"].get((cell["column"], cell["row"])) or []
            formatted_images = []
            for img in images:
                styles = render_inline_styles(img["style"])
                img_tag = (
                    '<img width="{width}" height="{height}"'
                    'style="{styles_str}"'
                    'src="{src}"'
                    "/>"
                ).format(styles_str=styles, **img)
                formatted_images.append(img_tag)
            trow.append(
                (
                    '<td {attrs_str} style="{styles_str}">'
                    "{formatted_images}"
                    "{formatted_value}"
                    "</td>"
                ).format(
                    attrs_str=render_attrs(cell["attrs"]),
                    styles_str=render_inline_styles(cell["style"]),
                    formatted_images="\n".join(formatted_images),
                    **cell,
                )
            )

        trow.append("</tr>")
        html.append("\n".join(trow))
    html.append("</table>")
    return "\n".join(html)


HTML_TEMPLATE = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <title>Title</title>
    </head>
    <body>
        %s
    </body>
    </html>
    """


def render_data_to_html(data, append_headers, append_lineno):
    return HTML_TEMPLATE % render_table(data, append_headers, append_lineno)


def get_sheet(wb, sheet):
    ws = wb.active
    if sheet is not None:
        try:
            ws = wb[sheet]
        except KeyError:
            ws = wb.worksheets[sheet]
    return ws


def xlsx2html(
    filepath,
    output=None,
    locale="en",
    sheet=None,
    parse_formula=False,
    append_headers=(lambda dumb1, dumb2: True),
    append_lineno=(lambda dumb1, dumb2: True),
    default_cell_border="none",
):
    """

    :param filepath: the path to open or a file-like object
    :type filepath: string or a file-like object open in binary mode c.f., :class:`zipfile.ZipFile`
    :param output: the path to open or a file-like object
    :param locale: string or a file-like object open in binary mode c.f., :class:`zipfile.ZipFile`
    :param sheet: if `None` - first sheet; if `-1` - all sheets;
        if string sheet name.
        if number - sheet index. can also use list of names or indexes
    :param parse_formula:
    :param append_headers:
    :param append_lineno:
    :param default_cell_border:
    :return:
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_list = [sheet]
    if isinstance(sheet, (list, tuple)):
        # TODO any iterable
        sheet_list = sheet
    elif sheet == -1:
        sheet_list = wb.sheetnames

    if not output:
        output = io.StringIO()
    if isinstance(output, str):
        output = open(output, "w", encoding="utf-8")
    if output.encoding and output.encoding not in ["utf-8", "utf-16"]:
        raise UnicodeError("output must be opened with encoding='utf-8'")

    html_tables = []
    for sheet in sheet_list:
        ws = get_sheet(wb, sheet)
        fs = None
        if parse_formula:
            fb = openpyxl.load_workbook(filepath, data_only=False)
            fs = get_sheet(fb, sheet)

        data = worksheet_to_data(
            ws, locale=locale, fs=fs, default_cell_border=default_cell_border
        )
        html_tables.append(render_table(data, append_headers, append_lineno))

    html = HTML_TEMPLATE % "\n".join(html_tables)
    output.write(html)
    output.flush()
    return output
