# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import openpyxl
import sys


def worksheet_to_data(ws):
    merged_cell_map = {}
    exclded_cells = set(ws.merged_cells)
    for cell_range in ws.merged_cell_ranges:
        cell_range_list = list(ws[cell_range])
        m_cell = cell_range_list[0][0]
        merged_cell_map[m_cell.coordinate] = {
            'colspan': len(cell_range_list[0]),
            'rowspan': len(cell_range_list),
        }
        exclded_cells.remove(m_cell.coordinate)

    data_list = []
    for row in ws.iter_rows():
        data_row = []
        data_list.append(data_row)
        for cell in row:
            if cell.coordinate in exclded_cells:
                continue
            cell_data = {
                'value': cell.value or '',
                'colspan': 1,
                'rowspan': 1,
            }
            cell_data.update(merged_cell_map.get(cell.coordinate) or {})
            data_row.append(cell_data)
    return data_list


def render_data_to_html(data):
    html = ['<html>'
            '<head>'
            '<meta charset="utf-8">'
            '</head>'
            '<body>', '<table>']
    for row in data:
        trow = ['<tr>']
        for col in row:
            trow.append('<td colspan="{colspan}" rowspan="{rowspan}">{value}</td>'.format(**col))
        trow.append('</tr>')
        html.append('\n'.join(trow))
    html.append('</table></body></html>')
    return '\n'.join(html)


def xls2html(filepath, output):
    ws = openpyxl.load_workbook(filepath).active
    data = worksheet_to_data(ws)
    html = render_data_to_html(data)

    with open(output, 'wb') as f:
        f.write(str(html.encode('utf-8')))



if __name__ == '__main__':
    xls2html(sys.argv[1], sys.argv[2])
