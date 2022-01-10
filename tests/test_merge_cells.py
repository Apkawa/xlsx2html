import pytest
from openpyxl.utils import range_boundaries
from xlsx2html import xlsx2html
from xlsx2html.parser.parser import XLSXParser


def test_monkey_patched():
    assert range_boundaries._monkey


@pytest.mark.parametrize(
    "range,expect_result",
    [
        ["A:A", (1, 1, 1, 1048576)],
        ["2:2", (1, 2, 18278, 2)],
        ["B2:B", (2, 2, 2, 1048576)],
        ["D2:2", (4, 2, 18278, 2)],
    ],
)
def test_range_boundaries(range, expect_result):
    result = range_boundaries(range)
    assert result == expect_result


@pytest.mark.skip("TODO optimize large merges")
def test_parse_merge_cells(fixture_file):
    p = XLSXParser(filepath=fixture_file("merge_cells.xlsx"))
    result = p.get_sheet()
    assert result
    assert len(result.cols) == 4
    assert len(result.rows) == 11


@pytest.mark.skip("TODO optimize large merges")
@pytest.mark.webtest()
def test_merge_cells(temp_file, browser, screenshot_regression, fixture_file):
    browser.driver.set_window_size(1280, 1024)
    out_file = temp_file()

    xlsx2html(fixture_file("merge_cells.xlsx"), out_file, locale="en")

    browser.visit("file://" + out_file)
    screenshot_regression()
