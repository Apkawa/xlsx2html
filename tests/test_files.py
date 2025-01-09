# coding: utf-8
import io
import os
import time

import openpyxl
import pytest
from PIL import Image

from tests.conftest import IN_GITHUB_ACTIONS
from xlsx2html.core import xlsx2html, worksheet_to_data

FIXTURES_ROOT = os.path.join(os.path.dirname(__file__), "fixtures")


def get_fixture(name):
    return os.path.join(FIXTURES_ROOT, name)


XLSX_FILE = get_fixture("example.xlsx")


def test_xlsx2html(temp_file):
    out_file = temp_file()
    xlsx2html(XLSX_FILE, out_file, locale="en")
    result_html = open(out_file).read()
    assert result_html


def test_use_streams():
    xlsx_file = open(XLSX_FILE, "rb")
    out_file = io.StringIO()
    xlsx2html(xlsx_file, out_file, locale="en")
    assert out_file.tell() > 0
    out_file.seek(0)
    result_html = out_file.read()
    assert result_html


def test_default_output():
    xlsx_file = open(XLSX_FILE, "rb")
    out_file = xlsx2html(xlsx_file, locale="en")
    assert out_file.tell() > 0
    out_file.seek(0)
    result_html = out_file.read()
    assert result_html


def test_hyperlink(temp_file):
    out_file = temp_file()
    xlsx2html(get_fixture("hyperlinks.xlsx"), out_file, locale="en", parse_formula=True)
    result_html = open(out_file).read()
    assert result_html


def test_issue_30_cell_range_value(temp_file):
    out_file = temp_file()
    xlsx2html(
        get_fixture("cell_range_value.xlsx"), out_file, locale="en", parse_formula=True
    )
    result_html = open(out_file).read()
    assert result_html


def test_issue_43(temp_file):
    """
    For reproduce run test with non utf-8 locale
    1) add line "bg_BG.CP1251 CP1251" into `/etc/locale.gen`
    2) sudo run locale-gen
    3) LC_ALL=bg_BG.CP1251 pytest -k test_issue_43
    """

    out_file = temp_file()
    xlsx2html(get_fixture("fileoutpart12.xlsx"), out_file, parse_formula=True)
    result_html = open(out_file).read()
    assert result_html

    stream = xlsx2html(
        get_fixture("fileoutpart12.xlsx"), output=None, parse_formula=True
    )
    stream.seek(0)
    result_html = stream.read()
    assert result_html

    with pytest.raises(UnicodeError) as e:
        # simulate open system encoding
        out_file = open(temp_file(), "w", encoding="cp1251")
        xlsx2html(get_fixture("fileoutpart12.xlsx"), out_file, parse_formula=True)
    assert "output must be opened with encoding='utf-8'" == str(e.value)


def test_issue_x000D(temp_file):
    """
    Fix _x000D_ issue
    https://stackoverflow.com/questions/29976234/openpyxl-unicode-values
    """
    wb = openpyxl.load_workbook(get_fixture("fileoutpart12.xlsx"))
    ws = wb.worksheets[0]
    cell = ws["A1"]
    # opps :3
    assert cell.value == "Parameter _x000D_"
    data = worksheet_to_data(ws)
    data_cell = data["rows"][0][0]
    assert data_cell["value"] == "Parameter \r"
    assert data_cell["formatted_value"] == "Parameter \r"


def page_has_loaded(self):
    # via https://stackoverflow.com/questions/26566799/wait-until-page-is-loaded-with-selenium-webdriver-for-python
    page_state = self.driver.execute_script("return document.readyState;")
    return page_state == "complete"


@pytest.mark.webtest()
@pytest.mark.skipif(
    IN_GITHUB_ACTIONS,
    reason="Have issue with no match screenshot size same screen size",
)
def test_screenshot_diff(temp_file, browser, screenshot_regression):
    browser.driver.set_window_size(1280, 1024)
    out_file = temp_file()
    xlsx2html(XLSX_FILE, out_file, locale="en")
    browser.visit("file://" + out_file)
    # Wait loading page
    from selenium.webdriver.support.wait import WebDriverWait

    WebDriverWait(browser, timeout=10).until(page_has_loaded)
    time.sleep(1)
    # Debug CI
    print("Window size", browser.driver.get_window_size())
    screenshot_file = temp_file(extension=".png")
    browser.driver.save_screenshot(screenshot_file)
    im = Image.open(screenshot_file)
    assert im.size == (1280, 1024)
    # End debug CI
    screenshot_regression()


def test_multiple_sheets(temp_file):
    out_file = temp_file()
    xlsx2html(XLSX_FILE, out_file, locale="en", sheet=[0, 2])  # Can use index
    result_html = open(out_file).read()
    assert result_html.count("</table>") == 2

    # By names
    wb = openpyxl.load_workbook(XLSX_FILE)
    xlsx2html(XLSX_FILE, out_file, locale="en", sheet=wb.sheetnames)  # Can use index
    result_html = open(out_file).read()
    assert result_html.count("</table>") == 3

    out_file = temp_file()
    xlsx2html(XLSX_FILE, out_file, locale="en", sheet=-1)  # All sheets
    result_html = open(out_file).read()
    assert result_html.count("</table>") == 3
